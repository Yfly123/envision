import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def Textcolor(wk,st,en,flag):
    sheet1 = wk['Sheet1']#wk[wk_name[0]]#title名称
    en=st+1
    for i in range(st,en):
        if flag ==0:
            Color = ['FFFFFF']
        elif flag==1:
            Color = ['FFCC66']# 黄
        elif flag==2:
            Color = ['FF9966']#橙
        elif flag==3:
            Color = ['FF3333']#浅红
        else:
            Color = ['990000']#深红
        fille = PatternFill('solid', fgColor=Color[0])  # 设置填充颜色为 橙色
        sheet1.cell(row=i+1 , column=2).fill = fille  # 序列
file_addr = r'C:\Users\fei.yang4\Desktop\tiaoshui_test-4.xlsx'
wk = openpyxl.load_workbook(file_addr)  # 加载已经存在的excel
data = pd.read_excel(file_addr)
df =pd.DataFrame(data)
flag = 0
t_temp=0
i,j=0,50

while i< len(df['rg']) and len(df['rg'])-i>j-i:
    if flag==0:
        j=i+50
        if abs((df.iat[i,1]-df.iat[j,1]))*100>=0.8:
            flag=1
            t_temp=1
            Textcolor(wk, i + 1+50, j + 1, flag)
            print(abs((df.iat[i,1]-df.iat[j,1]))*100)
            print('i=%s, j=%s' % (i, j))
            continue

    if flag==1:
        if t_temp==1:
            i = j
            j=i+20
            t_temp=2
        else:
            j=i+20
            if abs((df.iat[i,1]-df.iat[j,1]))*100>=0.25:
                flag=2
                Textcolor(wk,i+1+20, j+1, flag)
                print(abs((df.iat[i, 1] - df.iat[j, 1])) * 100)
                print('i=%s, j=%s' % (i, j))
                continue

    if flag == 2:
        if t_temp==2:
            i = j
            j = i+10
            t_temp=3
        else:
            j=i+10
            # Textcolor(wk, i + 1, j + 1, flag)
            if abs((df.iat[i,1]-df.iat[j,1]))*100>=0.16:
                flag=3
                Textcolor(wk, i + 1+10, j + 1, flag)
                print('i=%s, j=%s' % (i, j))
                print(abs((df.iat[i, 1] - df.iat[j, 1])) * 100)
                continue
    if flag ==3:
        if t_temp==3:
            i = j
            j = i+5
            t_temp=4
        else:
            j =i +5
            if abs((df.iat[i, 1] - df.iat[j, 1]))*100 >= 0.1:
                flag=4
                Textcolor(wk, i+1+5, j+1, flag)
                print(abs((df.iat[i, 1] - df.iat[j, 1])) * 100)
                # print(abs((df.iat[i, 1] - df.iat[j, 1])) * 100)
                break
            print('i=%s, j=%s' % (i, j))
    i +=1
wk.save(file_addr)  # 保存excel
